"""
Módulo de exportación a Excel y PDF
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import tkinter as tk
from tkinter import filedialog, messagebox
from datetime import datetime
import os


class Exporters:

    @staticmethod
    def exportar_a_excel(datos, columnas, titulo="Reporte"):
        """Exportar datos a Excel"""
        try:
            # Seleccionar ubicación
            filename = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"{titulo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )

            if not filename:
                return

            # Crear libro y hoja
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = titulo

            # Estilos
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Escribir encabezados
            for col, header in enumerate(columnas, 1):
                cell = ws.cell(row=1, column=col)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
                cell.border = border

            # Escribir datos
            for row_idx, row_data in enumerate(datos, 2):
                for col_idx, col_name in enumerate(columnas, 1):
                    # Buscar la clave en el diccionario (convertir a minúsculas y quitar espacios)
                    key = col_name.lower().replace(' ', '_')
                    valor = row_data.get(key, '')

                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = valor
                    cell.border = border
                    cell.alignment = Alignment(horizontal="left" if isinstance(valor, str) else "right")

                    # Formato para números
                    if isinstance(valor, (int, float)):
                        if key in ['precio', 'precio_compra', 'precio_venta', 'total', 'subtotal']:
                            cell.number_format = '#,##0.00'
                            cell.value = float(valor)

            # Ajustar ancho de columnas
            for col in ws.columns:
                max_length = 0
                col_letter = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[col_letter].width = adjusted_width

            # Guardar archivo
            wb.save(filename)
            messagebox.showinfo("Éxito", f"Archivo exportado exitosamente:\n{filename}")

        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar a Excel:\n{str(e)}")

    @staticmethod
    def exportar_a_pdf(datos, columnas, titulo="Reporte"):
        """Exportar datos a PDF"""
        try:
            filename = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                initialfile=f"{titulo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            )

            if not filename:
                return

            doc = SimpleDocTemplate(filename, pagesize=landscape(letter))
            elementos = []

            styles = getSampleStyleSheet()
            title_style = styles['Title']
            normal_style = styles['Normal']

            # Título
            titulo_text = Paragraph(f"<b>{titulo}</b>", title_style)
            elementos.append(titulo_text)
            elementos.append(Spacer(1, 0.2 * inch))

            # Fecha
            fecha_text = Paragraph(f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}", normal_style)
            elementos.append(fecha_text)
            elementos.append(Spacer(1, 0.2 * inch))

            # Preparar datos para tabla
            data = [columnas]
            for row in datos:
                fila = []
                for col in columnas:
                    key = col.lower().replace(' ', '_')
                    valor = row.get(key, '')
                    if isinstance(valor, (int, float)):
                        if 'precio' in key or 'total' in key or 'subtotal' in key:
                            fila.append(f"${valor:,.2f}")
                        else:
                            fila.append(str(valor))
                    else:
                        fila.append(str(valor)[:50] if valor else '')
                data.append(fila)

            table = Table(data, repeatRows=1)

            style = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#F8F9FA')),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ])

            for i in range(1, len(data)):
                if i % 2 == 0:
                    style.add('BACKGROUND', (0, i), (-1, i), colors.HexColor('#E9ECEF'))

            table.setStyle(style)
            elementos.append(table)

            elementos.append(Spacer(1, 0.2 * inch))
            pie = Paragraph(f"<i>Total de registros: {len(datos)}</i>", normal_style)
            elementos.append(pie)

            doc.build(elementos)
            messagebox.showinfo("Éxito", f"PDF exportado exitosamente:\n{filename}")

        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar a PDF:\n{str(e)}")